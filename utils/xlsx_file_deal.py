import datetime
from typing import Dict, List

from openpyxl import Workbook, load_workbook
import os.path

def file_parsing(file_name: str, max_rows: int, ws: str) -> List[List[str]]:

    workbook = load_workbook(file_name)
    # 获取某个工作表
    worksheet = workbook[ws]
    # 逐行读取工作表
    i = 0
    data_row: List[List[str]] = []
    for row in worksheet.iter_rows(min_row=1):
        # 遍历行中的单元格
        data_column: List[str] = []
        for cell in row:
            value = cell.value if cell.value else ""

            if value and isinstance(value, datetime.datetime):
                value = value.strftime('%Y-%m-%d')

            data_column.append(str(value))
        data_row.append(data_column)
        i += 1
        if i >= max_rows:
            yield data_row
            data_row = []
            i = 0
    yield data_row
    # 关闭工作簿
    workbook.close()


def file_append(file_name: str, ws: str,datas:[str]):

    if not os.path.exists(file_name):
        workbook = Workbook()
        workbook.active.title = ws
        workbook.save(filename=file_name)

    # 打开现有的Excel文件
    workbook = load_workbook(filename=file_name)

    # 获取指定的工作表
    worksheet = workbook[ws]

    # 追加新的数据到工作表中
    for data in datas:
        worksheet.append(data)

    # 保存Excel文件
    workbook.save(filename=file_name)

    # 关闭工作簿
    workbook.close()




if __name__ == '__main__':
    file_name = "example.xlsx"
    for data in file_parsing(file_name, 100,"Sheet1"):
        print(data)
